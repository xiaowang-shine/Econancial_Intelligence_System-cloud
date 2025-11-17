# async_file_processor.py - 异步文件处理模块
import os
import time
import threading
import queue
import pandas as pd
from typing import Dict, Any, Optional, Callable
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class AsyncFileProcessor:
    """异步文件处理器"""
    
    def __init__(self, max_workers: int = 4, chunk_size: int = 10000):
        """
        初始化异步文件处理器
        
        Args:
            max_workers: 最大工作线程数
            chunk_size: 文件分块大小
        """
        self.max_workers = max_workers
        self.chunk_size = chunk_size
        self.executor = ThreadPoolExecutor(max_workers=max_workers)
        self.task_queue = queue.Queue()
        self.active_tasks = {}
        self.task_results = {}
        
    def process_file_async(self, task_id: str, file_path: str, 
                          processing_func: Callable, **kwargs) -> str:
        """
        异步处理文件
        
        Args:
            task_id: 任务ID
            file_path: 文件路径
            processing_func: 处理函数
            **kwargs: 传递给处理函数的参数
            
        Returns:
            任务ID
        """
        # 提交异步任务
        future = self.executor.submit(
            self._process_file_with_progress,
            task_id, file_path, processing_func, **kwargs
        )
        
        self.active_tasks[task_id] = future
        logger.info(f"提交异步任务: {task_id}")
        
        return task_id
    
    def _process_file_with_progress(self, task_id: str, file_path: str,
                                   processing_func: Callable, **kwargs) -> Dict[str, Any]:
        """
        带进度的文件处理
        
        Args:
            task_id: 任务ID
            file_path: 文件路径
            processing_func: 处理函数
            **kwargs: 传递给处理函数的参数
            
        Returns:
            处理结果
        """
        try:
            logger.info(f"开始处理文件: {file_path}")
            start_time = time.time()
            
            # 获取文件大小用于进度计算
            file_size = os.path.getsize(file_path)
            
            # 分块读取和处理文件
            result = self._process_large_file(file_path, processing_func, **kwargs)
            
            processing_time = time.time() - start_time
            logger.info(f"文件处理完成: {task_id}, 耗时: {processing_time:.2f}秒")
            
            result['processing_time'] = processing_time
            result['file_size'] = file_size
            
            self.task_results[task_id] = result
            return result
            
        except Exception as e:
            logger.error(f"文件处理失败: {task_id}, 错误: {str(e)}")
            error_result = {'error': str(e), 'task_id': task_id}
            self.task_results[task_id] = error_result
            return error_result
    
    def _process_large_file(self, file_path: str, processing_func: Callable, **kwargs) -> Dict[str, Any]:
        """
        分块处理大文件
        
        Args:
            file_path: 文件路径
            processing_func: 处理函数
            **kwargs: 传递给处理函数的参数
            
        Returns:
            处理结果
        """
        try:
            # 尝试分块读取Excel文件
            if file_path.endswith(('.xlsx', '.xls')):
                return self._process_excel_chunks(file_path, processing_func, **kwargs)
            else:
                # 其他文件类型的处理
                return processing_func(file_path, **kwargs)
                
        except Exception as e:
            logger.error(f"文件处理错误: {str(e)}")
            raise
    
    def _process_excel_chunks(self, file_path: str, processing_func: Callable, **kwargs) -> Dict[str, Any]:
        """
        分块处理Excel文件
        
        Args:
            file_path: 文件路径
            processing_func: 处理函数
            **kwargs: 传递给处理函数的参数
            
        Returns:
            处理结果
        """
        # 首先检查文件大小
        file_size = os.path.getsize(file_path)
        
        # 如果文件小于10MB，直接处理
        if file_size < 10 * 1024 * 1024:
            df = pd.read_excel(file_path)
            return processing_func(df, **kwargs)
        
        # 大文件分块处理
        logger.info(f"处理大文件: {file_path}, 大小: {file_size / (1024*1024):.2f}MB")
        
        # 使用openpyxl分块读取
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filename=file_path, read_only=True)
            sheet = wb.active
            
            # 读取表头
            headers = [cell.value for cell in sheet[1]]
            
            # 分块读取数据
            chunks = []
            current_chunk = []
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                current_chunk.append(row)
                
                if len(current_chunk) >= self.chunk_size:
                    chunk_df = pd.DataFrame(current_chunk, columns=headers)
                    chunks.append(chunk_df)
                    current_chunk = []
            
            # 处理最后一个块
            if current_chunk:
                chunk_df = pd.DataFrame(current_chunk, columns=headers)
                chunks.append(chunk_df)
            
            wb.close()
            
            # 合并所有块的处理结果
            results = []
            for i, chunk in enumerate(chunks):
                logger.info(f"处理块 {i+1}/{len(chunks)}, 大小: {len(chunk)} 行")
                chunk_result = processing_func(chunk, **kwargs)
                results.append(chunk_result)
            
            # 合并结果
            if results:
                # 根据结果类型进行合并
                if isinstance(results[0], dict):
                    merged_result = self._merge_dict_results(results)
                elif isinstance(results[0], pd.DataFrame):
                    merged_result = pd.concat(results, ignore_index=True)
                else:
                    merged_result = results
                
                return {'result': merged_result, 'chunks_processed': len(chunks)}
            
        except ImportError:
            # 如果没有openpyxl，回退到普通读取
            logger.warning("未安装openpyxl，使用普通方式读取Excel文件")
            df = pd.read_excel(file_path)
            return processing_func(df, **kwargs)
    
    def _merge_dict_results(self, results: list) -> Dict[str, Any]:
        """
        合并字典类型的结果
        
        Args:
            results: 结果列表
            
        Returns:
            合并后的结果
        """
        merged = {}
        for result in results:
            if isinstance(result, dict):
                for key, value in result.items():
                    if key not in merged:
                        merged[key] = value
                    elif isinstance(value, list) and isinstance(merged[key], list):
                        merged[key].extend(value)
                    elif isinstance(value, dict) and isinstance(merged[key], dict):
                        merged[key].update(value)
        
        return merged
    
    def get_task_status(self, task_id: str) -> Dict[str, Any]:
        """
        获取任务状态
        
        Args:
            task_id: 任务ID
            
        Returns:
            任务状态信息
        """
        if task_id in self.active_tasks:
            future = self.active_tasks[task_id]
            if future.done():
                try:
                    result = future.result()
                    del self.active_tasks[task_id]
                    return {
                        'status': 'completed',
                        'result': result
                    }
                except Exception as e:
                    del self.active_tasks[task_id]
                    return {
                        'status': 'failed',
                        'error': str(e)
                    }
            else:
                return {'status': 'running'}
        elif task_id in self.task_results:
            return {
                'status': 'completed',
                'result': self.task_results[task_id]
            }
        else:
            return {'status': 'not_found'}
    
    def cancel_task(self, task_id: str) -> bool:
        """
        取消任务
        
        Args:
            task_id: 任务ID
            
        Returns:
            是否成功取消
        """
        if task_id in self.active_tasks:
            future = self.active_tasks[task_id]
            cancelled = future.cancel()
            if cancelled:
                del self.active_tasks[task_id]
            return cancelled
        return False
    
    def cleanup_completed_tasks(self):
        """清理已完成的任务"""
        completed_tasks = []
        for task_id, future in self.active_tasks.items():
            if future.done():
                completed_tasks.append(task_id)
        
        for task_id in completed_tasks:
            del self.active_tasks[task_id]
        
        logger.info(f"清理了 {len(completed_tasks)} 个已完成任务")
    
    def shutdown(self):
        """关闭处理器"""
        self.executor.shutdown(wait=True)
        logger.info("异步文件处理器已关闭")

# 全局实例
file_processor = AsyncFileProcessor()